import pandas
import requests
from lxml import html
from openpyxl import load_workbook
import os
import sys
import sqlite3
import time
from datetime import datetime
import multiprocessing
import pprint
import bz2
import json
import pickle
import re
import openpyxl
from urllib.parse import urljoin
from tqdm import tqdm

num_processes = multiprocessing.cpu_count()


def read_excel():
    path = "AzureMachine/employee_file (1).xlsx"
    df = pandas.read_excel(path, engine='openpyxl')
    return df


def buildregex():
    wb = openpyxl.load_workbook("Keywords for new project.xlsx")
    ws = wb.active
    regex = ""
    used = []
    for row in ws.iter_rows(values_only=True):
        if row[0] in used:
            continue
        else:
            used.append(row[0])
            regex = regex + str(row[0]) + "|"  # f"|(>\s*\S*\s*{row[0]}\s*(<BR>)?\s*(</[a-zA-Z]*>)+)"
    regex = regex[:-1]
    return regex


def process(df):
    numbers = []
    percents = []
    certains = []
    for idx, row in tqdm(df.iterrows(), total=df.shape[0]):
        number = None
        percent = None
        certain = None

        # sentence = str(row['Keyword Sentence']).replace(',', '').lower()
        sentence = str(row['Employees']).replace(',', '').lower().encode('ascii', errors='ignore').decode('ascii')
        #print(sentence)
        if ('("eu")') in sentence or "european" in sentence:
            pass
        else:
            ns = None
            if "of our" in sentence:
                ns = re.findall(r'(\d+) of our', sentence)
                if ns:
                    number = ns[0]
            if not ns:
                ns = re.findall(rf'(?:([0-9]*.?[0-9]+?%?)[a-zA-Z,\-%()\"\/| ]*(?:{buildregex()}))|(?:{buildregex()})[a-zA-Z,%()\"\/| ]*(\d+.?\d*)', sentence, re.IGNORECASE) #(rf'[\w\s\S]* (\d+.?\d*) [\w\s\S]*({buildregex()})', sentence, re.IGNORECASE)
                if ns:
                    number = str(ns[0][0])
                    if any(c.isalpha() for c in number):  # Handling letters
                        number = None
                    if f'in {number} ' in sentence:
                        number = None
                    if '%' in str(number):
                        percent = number.replace('%', '').strip()
                        number = None
                    else:
                        number = str(number).replace('.', '').strip()
                    # if str(number) >= '2000' and str(number) <= '2020':
                    #     number = None

                # else:
                #     ns = re.findall(rf'({buildregex()})[\w\s\S]* (\d+.?\d*) [\w\s\S]*', sentence, re.IGNORECASE)
                #     if ns:
                #         number = ns[0][1]
                #     else:
                #         if 'various' in sentence or 'some of' in sentence or 'some or' in sentence:
                #             certain = 1

        if 'various' in sentence or 'some of' in sentence or 'some or' in sentence or 'certain of' in sentence or "Many of" in sentence:
            certain = 1
        nounion = ["not represented by", "none ", "not subject to", 'non-union', 'no unions', "do not cover any", "party to any", 'not unionized', "not covered by", "do not have "
                                                                                                     "employees that "
                                                                                                     "are represented "
                                                                                                     "by"]
        if any(word for word in nounion if word in sentence):
            number = 0
            percent = None
        # ns = re.findall(rf'[\S\s\w]*?(\d+.?\d*)(%| percent)[\w\s\S]*({buildregex()})', sentence, re.IGNORECASE)
        # if ns:
        #     percent = ns[0][0]
        #     if number:
        #         n1 = sentence.find(percent)
        #         n2 = sentence.find(number)
        #         if n1 > n2:
        #             number = None
        #         else:
        #             percent = None
        # else:
        #     percent = None
        numbers.append(number)
        percents.append(percent)
        certains.append(certain)
        # if idx == 57:
        #     print(number, percent, sentence)
        # if idx == 60:
        #     break
    df["Number"] = numbers
    df["Percent"] = percents
    df["Certain"] = certains
    return df

def write_df(df):
    df.to_excel("Numbers.xlsx", engine='openpyxl', index=False)


if __name__ == '__main__':
    write_df(process(read_excel()))