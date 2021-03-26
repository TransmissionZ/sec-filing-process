INPUT_FOLDER_NAME = "CIK/"


import re
import os
from bs4 import BeautifulSoup
import csv
import openpyxl
import lxml.html


class Parser:
    def __init__(self, input_folder):
        self.extracted = False
        self.input_folder = input_folder
        return

    def listdir(self, location):
        return os.listdir(self.input_folder + location + "/")

    def getpath(self, location, t):
        return os.path.join(self.input_folder, location, t)

    def parse(self):
        print("Parsing Text Files... ")
        c = 0
        wb = openpyxl.load_workbook("CIK.xlsx")
        ws = wb.active
        CIKs = []
        for row in ws.iter_rows(values_only=True):
            CIKs.append(row[0])

        with open('employee_file.csv', mode='w') as employee_file:
            employee_writer = csv.writer(employee_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            employee_writer.writerow(["CIK", "Filed As Of Date", "Conformed Period Of Report", "Keyword Whole", "Keyword Item 1"])
            l = len(os.listdir(self.input_folder))
            for p in os.listdir(self.input_folder):
                if len(self.listdir(p)) == 0:
                    continue
                if int(p) in CIKs:
                    c += 1
                    # print(p)
                    for t in (self.listdir(p)):
                        try:
                            self.extract(p, t, employee_writer)
                        except Exception as e:
                            print(e)
                if (c % 10 == 0):
                    print("CIK's Done: " + str(c) + "/" + str(l))

    def buildregex(self):
        wb = openpyxl.load_workbook("Keywords for new project.xlsx")
        ws = wb.active
        regex = ""
        used = []
        for row in ws.iter_rows(values_only=True):
            if row[0] in used:
                continue
            else:
                used.append(row[0])
                regex = regex + str(row[0]) + "|"  #f"|(>\s*\S*\s*{row[0]}\s*(<BR>)?\s*(</[a-zA-Z]*>)+)"
        regex = regex[:-1]
        return regex


    def extract(self, p, t, employee_writer=None):
        with open(self.getpath(p, t), 'r') as f:
            lines = f.readlines()
            for l in lines:
                if ("central index key:" in l.lower()):
                    n = l.find(":")
                    cik = l[n + 1:].strip()
                if ("conformed period of report:" in l.lower()):
                    n = l.find(":")
                    cpor = l[n + 1:].strip()
                if ("filed as of date:" in l.lower()):
                    n = l.find(":")
                    faod = l[n + 1:].strip()
            f.seek(0)
            page = f.read()

            # Pre-processing the html content by removing extra white space and combining then into one line.
            page = page.strip()  # <=== remove white space at the beginning and end
            page = page.replace('\n', ' ')  # <===replace the \n (new line) character with space
            page = page.replace('\r', '')  # <===replace the \r (carriage returns -if you're on windows) with space
            page = page.replace('&amp;', 'and')  # <===replace the \r (carriage returns -if you're on windows) with space
            page = page.replace('&#xA0;', ' ')  # <===replace the \r (carriage returns -if you're on windows) with space
            page = page.replace('&#160;', ' ')  # <===replace the \r (carriage returns -if you're on windows) with space
            page = page.replace('&#150;', ' ')  # <===replace the \r (carriage returns -if you're on windows) with space
            page = page.replace('&#151;', ' ')  # <===replace the \r (carriage returns -if you're on windows) with space
            page = page.replace('&#8211;', ' ')  # <===replace the \r (carriage returns -if you're on windows) with space
            page = page.replace('&#x2013;', ' ')  # <===replace the \r (carriage returns -if you're on windows) with space
            page = page.replace('&nbsp;',
                                ' ')  # <===replace "&nbsp;" (a special character for space in HTML) with space.
            page = page.replace('&#160;',
                                ' ')  # <===replace "&#160;" (a special character for space in HTML) with space.
            while '  ' in page:
                page = page.replace('  ', ' ')  # <===remove extra space

            # Using regular expression to extract texts that match a pattern

            # Define pattern for regular expression.
            # The following patterns find ITEM 1 and ITEM 1A as diplayed as subtitles
            # (.+?) represents everything between the two subtitles
            # If you want to extract something else, here is what you should change

            # Define a list of potential patterns to find ITEM 1 and ITEM 1A as subtitles
            regexs = (#r"(Item 1.)[\s\S]*(item 2.)", '')
                      #   'bold;\">\s*Item 1\.(.+?)bold;\">\s*Item 1A\.[\s\S]*(item 2.)',
                      # # <===pattern 1: with an attribute bold before the item subtitle
                      # 'b>\s*Item 1\.(.+?)b>\s*Item 1A\.[\s\S]*(item 2.)',  # <===pattern 2: with a tag <b> before the item subtitle
                      # 'Item 1\.\s*<\/b>(.+?)Item 1A\.\s*<\/b>[\s\S]*(item 2.)',
                      # # <===pattern 3: with a tag <\b> after the item subtitle
                      # 'Item 1\.\s*Business\.\s*<\/b(.+?)Item 1A\.\s*Risk Factors\.\s*<\/b[\s\S]*(item 2.)', # <===pattern 4: with a tag <\b> after the item+description subtitle
                        'Item 1\.\s*Business(\.)?\s*</b>[\s\S]*(item 2.)',
                        'Item 1\.\s*</b>[\s\S]*(item 2.)',
                        'Item 1\.\s*Business[\s\S]*(item 2.)',
                        'Items 1(\.)? (and|&) 2(\.|\:)[\s\S]*(item 3)',
                        'ITEMS 1., 1A. and 2.[\s\S]*(item 3)',
                        'Item 1\.\s*</a>[\s\S]*(item 2.)',
                        'Item 1\.\s*<(/)?font[\s\S]*(item 2.)',
                        'Item 1\.\s*</span><span[\s\S]*(item 2.)',
                        'Item 1\.\s*B[\s\S]*(item 2.)',
                        'Item 1(\.)?\s*Description of Business<[\s\S]*(item 2)',
                        'PART I\s*</font>[\s\S]*(item 2.)',
                        'Item 1\:\s*Business[\s\S]*(item 2:)',
                        'Item 1(\.)?</u>[\s\S]*(item 2)',
                        '1\.\s*Business[\s\S]*(2. Properties)',
                        'tem 1\.\s*Business[\s\S]*(tem 2.)',
                        'Item 1<a[\s\S]*\:\s*Business[\s\S]*(item 2)',
                        'Item 1(\.)?\s*<(/)?b>[\s\S]*(item 2(\.)?)',
                        'ITEM 1<BR> BUSINESS[\s\S]*(item 2)',
                        'ITEM 1.</TD>[\s\S]*(item 2.</TD>)',
                        'ITEM (</a>)?1\s*-\s*BUSINESS[\s\S]*(item (</a>)?2)',
                        'Item 1\:\s* <A[\s\S]*(item 2:)',
                        'Item 1\.[\s\S]*(item 2.)',
                        'item 1\s*</font>[\s\S]*(item 2</FONT>)',
                        'Item 1\s*Business<[\s\S]*(item 2\s*Properties)',
                        'Item No. 1\s*(-)?\s*Description of Business<[\s\S]*(Item No. 2\s*(-)?\s*Properties<)',
                        '>Business <[\s\S]*(>Properties <)',
                        'ITEMS 1, 1A, and 2<br>[\s\S]*(item 3)',
                        'Item 1\s*Business[\s\S]*(item 2\s*properties)',
                        'Item 1\:[\s\S]*(item 2:)',
                        'Item 1\s*Busines[\s\S]*(item 2\s* Properties)',
                        'tem 1\.[\s\S]*(tem 2.)',
                        'item </a>1.[\s\S]*(item </a>2.)',
                        'Item 1[\s\S]Item 2'
                      )

            # Now we try to see if a match can be found...
            for regex in regexs:
                match = re.search(regex, page, flags=re.IGNORECASE)  # <===search for the pattern in HTML using re.search from the re package. Ignore cases.
                # print(match.group())
                # If a match exist....
                if match:
                    # print("matched + " + str(p) + " " + str(t))
                    soup = BeautifulSoup(match.group(), "lxml")  # <=== match.group(1) returns the texts inside the parentheses (.*?)

                    # soup.text removes the html tags and only keep the texts
                    rawText = soup.text # <=== you have to change the encoding the unicodes
                    # remove space at the beginning and end and the subtitle "business" at the beginning
                    # ^ matches the beginning of the text
                    # outText = re.sub("^business\s*", "", rawText.strip(), flags=re.IGNORECASE)

                    #print(rawText.strip()[:100])
                    # output_file = open(output_path, "w")
                    # output_file.write(outText)
                    # output_file.close()

                    break  # <=== if a match is found, we break the for loop. Otherwise the for loop continues

            keywords = self.buildregex()
            c4 = re.search(keywords, page)
            if c4:
                c4 = 1
            else:
                c4 = 0
            c5 = 0
            try:
                if rawText and c4 != 0:
                    c5 = re.search(keywords, rawText)
                    if c5:
                        c5 = 1
                    else:
                        c5 = 0
                else:
                    c5 = 0
            except:
                c5 = ""
                print("not-matched + " + str(p) + " " + str(t))
            # print(c5)
            employee_writer.writerow([cik, faod, cpor, c4, c5])


if __name__ == '__main__':
    p = Parser(INPUT_FOLDER_NAME)
    # p.buildregex()
    # p.extract('3545', '0000003545-15-000102.txt')
    p.parse()