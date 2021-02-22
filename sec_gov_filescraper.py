INPUT_TYPES = ["10-K"]
INPUT_START_YEAR = 2009
INPUT_END_YEAR = 2019
INPUT_FILE = "CIK.xlsx"
INPUT_SHEET = "Sheet1"
INPUT_FOLDER_NAME = "CIK"


# other inputs, should be the same
INPUT_DATABASE_NAME = "SecFileDatabase.db"
INPUT_BATCH_SIZE = 50 # how many parallel downloads


# install: requests, lxml, openpyxl
import requests
from lxml import html
from openpyxl import load_workbook
import os
import sys
import sqlite3
import time
from datetime import datetime
import threading
import pprint
import bz2
import json
import pickle
from urllib.parse import urljoin



class Scraper:
    def __init__(self, input_file, input_sheet, input_types, input_start_year, input_end_year, input_database_name, input_batchsize, input_folder):
        ## check inputs first
        self.is_interrupted = False
        self.inputs_are_good = True

        if type(input_file) != str:
            print("Input file must be a string!")
            self.inputs_are_good = False
        if type(input_sheet) != str:
            print("Input sheet must be a string!")
            self.inputs_are_good = False

        if type(input_types) != list:
            print("Input types must be a list of strings!")
            self.inputs_are_good = False
        else:
            for probe_type in input_types:
                if type(probe_type) != str:
                    print("Input types must be a list of strings!")
                    self.inputs_are_good = False
                    break

        if type(input_start_year) != int:
            print("Start year must be an integer!")
            self.inputs_are_good = False
        if type(input_end_year) != int:
            print("End year must be an integer!")
            self.inputs_are_good = False

        if type(input_start_year) == int and type(input_end_year) == int:
            if input_start_year > input_end_year:
                print("Start year has to be <= end year!")
                self.inputs_are_good = False

        if type(input_database_name) != str:
            print("Database name must be a string!")
            self.inputs_are_good = False

        if type(input_batchsize) != int:
            print("Batch size must be an integer!")
            self.inputs_are_good = False
        else:
            if input_batchsize <= 0:
                print("Batch size must be a positive integer!")
                self.inputs_are_good = False

        if type(input_folder) != str:
            print("Input folder must be a string!")
            self.inputs_are_good = False

        if self.inputs_are_good == False:
            print("Bad inputs!")
            return

        ## if still here, inputs are good so set them
        self.input_file = input_file
        self.input_sheet = input_sheet
        self.input_types = input_types
        self.start_year = input_start_year
        self.end_year = input_end_year
        self.database_name = input_database_name
        self.batch_size = input_batchsize
        self.input_folder = input_folder

        ## create database
        self.db_conn = sqlite3.connect(self.database_name, check_same_thread=False)
        self.db_cursor = self.db_conn.cursor()
        self.db_cursor.execute("CREATE TABLE IF NOT EXISTS ScrapedPaginationItems (start_year INTEGER NOT NULL, end_year INTEGER NOT NULL, cik INTEGER NOT NULL, number_of_items INTEGER, PRIMARY KEY(start_year, end_year, cik))")
        self.db_cursor.execute("CREATE TABLE IF NOT EXISTS FilesTable (docurl TEXT NOT NULL PRIMARY KEY, cik INTEGER, filing_type TEXT, filing_year INTEGER, filing_date TEXT, filing_stamp REAL, filename TEXT, fileurl TEXT)")

        ## initialize items for threading
        self.good_count = 0
        self.all_thread_items = {}
        self.LOCK = threading.Lock()

        ## read inputs
        self.codes_to_scrape = self.read_inputs()
        return


    def read_inputs(self):
        try:
            items_to_return = []
            input_wb = load_workbook(self.input_file)
            input_ws = input_wb[self.input_sheet]
            for current_row_number in range(2, input_ws.max_row+1):
                potential_ciknumber = input_ws.cell(row=current_row_number, column=1).value
                try:
                    probe_cik = int(potential_ciknumber)
                    items_to_return.append({"cikint":probe_cik})
                except ValueError:
                    print("Can't read an integer in row", current_row_number, "in", self.input_file)
                    continue

            return items_to_return
        except:
            print("An exception while reading inputs - make sure filename and sheetname are correct!")
            return []

    
    def get_doc_links(self):
        if self.inputs_are_good == False or self.is_interrupted == True:
            return

        print("Getting document links from pagination...")
        print("Total codes loaded:", len(self.codes_to_scrape))
        ## scrape each unscraped code
        for code_index, code_to_scrape in enumerate(self.codes_to_scrape):
            existence_check = self.db_cursor.execute("SELECT EXISTS(SELECT 1 FROM ScrapedPaginationItems WHERE start_year=? AND end_year=? AND cik=?)", (self.start_year, self.end_year, code_to_scrape["cikint"])).fetchone()[0]
            if existence_check == 1:
                continue # already scraped this one!

            # if here, must scrape
            data_to_save = {}
            current_page = 1
            current_url = 'https://www.sec.gov/cgi-bin/browse-edgar?CIK=' + str(code_to_scrape["cikint"])
            pagination_scraped = False
            page_load_timeout = 60.0
            page_started_scraping_at = time.time()

            while 1:
                if time.time() - page_started_scraping_at >= page_load_timeout:
                    break # timeout

                need_to_break = False
                try:
                    r = requests.get(current_url, timeout=25)
                    tree = html.fromstring(r.text)
                    listing_els = tree.xpath("//table[@class='tableFile2']/tr/td/..")
                    if len(listing_els) == 0:
                        # maybe it is invalid parameter
                        invalid_param_el = tree.xpath("//center/h1[text()='Invalid parameter'] | //input[@type='button' and contains(@value, 'Previous') and @onclick]")
                        if len(invalid_param_el) != 0:
                            need_to_break = True
                        else:
                            print("Couldn't find any items at", current_url)
                            continue

                    for listing_el in listing_els:
                        this_listing = {"docurl":None, "date":None, "year":None, "stamp":None, "type":None}
                        
                        listing_url_el = listing_el.xpath("./td[2]/a[@href]")
                        if len(listing_url_el) != 0:
                            this_listing["docurl"] = urljoin('https://www.sec.gov/', listing_url_el[0].attrib["href"])
                            
                        listing_date_el = listing_el.xpath("./td[4]")
                        if len(listing_date_el) != 0:
                            date_string = self.fix_string(listing_date_el[0].text_content())
                            if ' ' in date_string:
                                date_string = date_string[0:date_string.find(" ")]
                            try:
                                date_object = datetime.strptime(date_string, "%Y-%m-%d")
                                this_listing["date"] = date_string
                                this_listing["year"] = int(date_object.year)
                                this_listing["stamp"] = self.get_timestamp_from_datetime_object(date_object)

                                # see if time condition is reached
                                if this_listing["year"] < self.start_year:
                                    need_to_break = True
                                    break
                            except ValueError:
                                pass
                            
                        listing_type_el = listing_el.xpath("./td[1]")
                        if len(listing_type_el) != 0:
                            this_listing["type"] = self.fix_string(listing_type_el[0].text_content())

                        if this_listing["docurl"] != None and this_listing["date"] != None and this_listing["type"] != None:
                            ## good info
                            if this_listing["year"] >= self.start_year and this_listing["year"] <= self.end_year:
                                if this_listing["docurl"] not in data_to_save:
                                    data_to_save[this_listing["docurl"]] = this_listing
                        else:
                            print("Couldn't add a listing at", current_url)

                    print("Scraped page", current_page, "at", code_to_scrape["cikint"], "(", code_index+1, "/", len(self.codes_to_scrape), ")")
                    # try to find next page
                    next_page_el = tree.xpath("//input[@type='button' and contains(@value, 'Next') and @onclick]")
                    if len(next_page_el) == 0:
                        need_to_break = True
                        
                    if need_to_break == True:
                        pagination_scraped = True
                        break
                    else:
                        # set up next page
                        urlpart = next_page_el[0].attrib["onclick"]
                        current_url = urljoin('https://www.sec.gov/', urlpart[urlpart.find("'")+1 : urlpart.rfind("'")])
                        current_page+=1
                        page_started_scraping_at = time.time()
                except KeyboardInterrupt:
                    print("Manual interrupt, quit!")
                    self.is_interrupted = True
                    return
                except:
                    print("An exception at", code_to_scrape["cikint"])
                    continue

            if pagination_scraped == False:
                print("Couldn't scrape pagination for", code_to_scrape["cikint"])
            else:
                # save results
                for item_to_save in data_to_save:
                    self.db_cursor.execute("INSERT OR IGNORE INTO FilesTable(docurl, cik, filing_type, filing_year, filing_date, filing_stamp) VALUES(?,?,?,?,?,?)",
                                           (data_to_save[item_to_save]["docurl"], code_to_scrape["cikint"], data_to_save[item_to_save]["type"], data_to_save[item_to_save]["year"],
                                            data_to_save[item_to_save]["date"], data_to_save[item_to_save]["stamp"]))
                self.db_cursor.execute("INSERT INTO ScrapedPaginationItems(start_year, end_year, cik, number_of_items) VALUES(?,?,?,?)",
                                       (self.start_year, self.end_year, code_to_scrape["cikint"], len(data_to_save)))
                self.db_conn.commit()
                print("Found", len(data_to_save), "items for", code_to_scrape["cikint"])
                    
            
        return


    def download_files(self):
        if self.inputs_are_good == False or self.is_interrupted == True:
            return

        # create main folder if it doesn't exist
        if not os.path.exists(self.input_folder):
            os.makedirs(self.input_folder)
            
        # get a list of links first, search for each of these where the file is
        doclinks_to_scrape = {}
        for input_type in self.input_types:
            this_type_fetcher = self.db_cursor.execute("SELECT docurl, cik FROM FilesTable WHERE filename IS NULL AND filing_type=? AND filing_year>=? AND filing_year<=?",
                                                       (input_type, self.start_year, self.end_year))
            for fetched_item in this_type_fetcher:
                if fetched_item[0] not in doclinks_to_scrape:
                    doclinks_to_scrape[fetched_item[0]] = {"docurl":fetched_item[0], "cik":fetched_item[1], "printnum":len(doclinks_to_scrape)+1, "folder":os.path.join(self.input_folder, str(fetched_item[1]))}

        print("Total documents left to scrape:", len(doclinks_to_scrape))

        # go scrape
        self.good_count = 0
        self.all_thread_items = {}
        for doc_to_scrape in doclinks_to_scrape:
            # create folder if it doesn't exist
            if not os.path.exists(doclinks_to_scrape[doc_to_scrape]["folder"]):
                os.makedirs(doclinks_to_scrape[doc_to_scrape]["folder"])
                
            self.all_thread_items[doc_to_scrape] = {"docurl":doclinks_to_scrape[doc_to_scrape]["docurl"], "cik":doclinks_to_scrape[doc_to_scrape]["cik"],
                                                    "printnum":doclinks_to_scrape[doc_to_scrape]["printnum"], "msg":None, "folder":doclinks_to_scrape[doc_to_scrape]["folder"]}
            if len(self.all_thread_items) == self.batch_size:
                ## call it
                all_threads = []
                for a_thread_item in self.all_thread_items:
                    current_thread = threading.Thread(target=self.download_thread, args=(self.all_thread_items[a_thread_item], ))
                    all_threads.append(current_thread)
                    current_thread.start()

                for thr in all_threads:
                    thr.join()

                for msg_to_check in self.all_thread_items:
                    if self.all_thread_items[msg_to_check]["msg"] != None:
                        print(self.all_thread_items[msg_to_check]["msg"])
                        
                print("Current item", doclinks_to_scrape[doc_to_scrape]["printnum"], "/", len(doclinks_to_scrape), "Good requests in this batch:", self.good_count, "/", len(self.all_thread_items))
                self.good_count = 0
                self.all_thread_items = {}

        if len(self.all_thread_items) != 0:
            ## call for residuals
            all_threads = []
            for a_thread_item in self.all_thread_items:
                current_thread = threading.Thread(target=self.download_thread, args=(self.all_thread_items[a_thread_item], ))
                all_threads.append(current_thread)
                current_thread.start()

            for thr in all_threads:
                thr.join()

            for msg_to_check in self.all_thread_items:
                if self.all_thread_items[msg_to_check]["msg"] != None:
                    print(self.all_thread_items[msg_to_check]["msg"])
                    
            print("Current item", doclinks_to_scrape[doc_to_scrape]["printnum"], "/", len(doclinks_to_scrape), "Good requests in this batch:", self.good_count, "/", len(self.all_thread_items))
            self.good_count = 0
            self.all_thread_items = {}
            
        return

    def download_thread(self, input_dict):
        ## find file url and name first
        file_url = None
        file_name = None
        
        try:
            first_req = requests.get(input_dict["docurl"], timeout=20)
            tree = html.fromstring(first_req.text)
            doc_el = tree.xpath("//p[text()='Document Format Files']/following-sibling::table[@class='tableFile'][1]//tr/td[text()='Complete submission text file']/following-sibling::td[1]/a[@href]")
            if len(doc_el) != 0:
                file_url = urljoin('https://www.sec.gov/', doc_el[0].attrib["href"])
                file_name = self.fix_string(doc_el[0].text_content()) #+ ".txt" #".bz2"
        except:
            return

        if file_url == None or file_name == None:
            self.all_thread_items[input_dict["docurl"]]["msg"] = "Couldn't find file element at " + input_dict["docurl"]
            return


        ## if still here, try to download it
        downloaded_fine = False
        try:
            file_req = requests.get(file_url, timeout=20)
            if file_req.status_code not in [200, 302]:
                self.all_thread_items[input_dict["docurl"]]["msg"] = "Status code not 200 or 302 at " + input_dict["docurl"]
                return

            #compressed_content = bz2.compress(file_req.content)
            with open(os.path.join(input_dict["folder"], file_name), 'wb') as save_fil:
                #save_fil.write(compressed_content)
                save_fil.write(file_req.content)
            downloaded_fine = True
        except:
            self.all_thread_items[input_dict["docurl"]]["msg"] = "An exception while requesting/parsing file at " + input_dict["docurl"]
            return

        # update database if good
        if downloaded_fine == True:
            with self.LOCK:
                try:
                    self.db_cursor.execute("UPDATE FilesTable SET filename=?, fileurl=? WHERE docurl=?", (file_name, file_url, input_dict["docurl"]))
                    self.db_conn.commit()
                    self.good_count+=1
                except:
                    pass
        
        return


    def fix_string(self, entry_string): # remove "\n", "\t" and double spaces
        exit_string = entry_string.replace("\n", "")
        exit_string = exit_string.replace("\t", "")
        exit_string = exit_string.replace("\r", "")
        while "  " in exit_string:
            exit_string = exit_string.replace("  ", " ")
        if len(exit_string) > 0: # remove first space
            if exit_string[0] == ' ':
                exit_string = exit_string[1:len(exit_string)]
        if len(exit_string) > 0: # remove last space
            if exit_string[len(exit_string)-1] == ' ':
                exit_string = exit_string[0:len(exit_string)-1]

        return exit_string

    def get_timestamp_from_datetime_object(self, input_object):
        # input_object must be in utc!!!!
        epoch = datetime(1970,1,1)
        time_diff_secs = (input_object - epoch).total_seconds()
        return time_diff_secs


if __name__ == '__main__':
    s = Scraper(INPUT_FILE, INPUT_SHEET, INPUT_TYPES, INPUT_START_YEAR, INPUT_END_YEAR, INPUT_DATABASE_NAME, INPUT_BATCH_SIZE, INPUT_FOLDER_NAME)
    s.get_doc_links()
    s.download_files()
