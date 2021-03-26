INPUT_FOLDER_NAME = "CIK/"


import re
import os
from bs4 import BeautifulSoup as soup
import csv
import openpyxl


class Parser:
    def __init__(self, input_folder):
        self.extracted = False
        self.input_folder = input_folder
        return

    def listdir(self, location):
        return os.listdir(self.input_folder + location + "/")

    def getpath(self, location, t):
        return os.path.join(self.input_folder, location, t)

    def parse(self):
        c = 0
        wb = openpyxl.load_workbook("CIK.xlsx")
        ws = wb.active
        CIKs = []
        for row in ws.iter_rows(values_only=True):
            CIKs.append(row[0])

        with open('employee_file.csv', mode='w') as employee_file:
            employee_writer = csv.writer(employee_file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            employee_writer.writerow(["CIK", "Filed As Of Date", "Conformed Period Of Report", "Employees"])
            l = len(os.listdir(self.input_folder))
            for p in os.listdir(self.input_folder):
                if len(self.listdir(p)) == 0:
                    continue
                if int(p) in CIKs:
                    c += 1
                    print(p)
                    for t in (self.listdir(p)):
                        try:
                            self.extract(p, t, employee_writer)
                        except Exception as e:
                            print(e)
                if (c % 10 == 0):
                    print("CIK's Done: " + str(c) + "/" + str(l))

    def buildregex(self):
        wb = openpyxl.load_workbook("Other Tittles.xlsx")
        ws = wb.active
        regex = "(>\s*\S*\s*(employees and Office Space|employees:?|employee relations:?"
        used = []
        for row in ws.iter_rows(values_only=True):
            if row[0] in used:
                continue
            else:
                used.append(row[0])
                regex = regex + "|" +str(row[0])  #f"|(>\s*\S*\s*{row[0]}\s*(<BR>)?\s*(</[a-zA-Z]*>)+)"
        regex += ")\s*(<BR>)?\s*(</[a-zA-Z]*>)+)"
        # print(regex)
        return regex

    def extract(self, p, t, employee_writer):
        employees = ""
        with open(self.getpath(p, t), 'r') as f:
            lines = f.readlines()
            for l in lines:
                if ("central index key:" in l.lower()):
                    n = l.find(":")
                    cik = l[n+1:].strip()
                if ("conformed period of report:" in l.lower()):
                    n = l.find(":")
                    cpor = l[n + 1:].strip()
                if ("filed as of date:" in l.lower()):
                    n = l.find(":")
                    faod = l[n + 1:].strip()
            try:
                # print(p + "/" + t)
                f.seek(0)
                file = f.read()
                n = [m.start() for m in re.finditer("<html>", file, re.IGNORECASE)] # file.lower().find("<html>")
                n1 = [m.end() for m in re.finditer("</html>", file, re.IGNORECASE)] # file.lower().find("</html>")
                # file = file[n[0]:n1[0]]

                for nn, nn1 in zip(n, n1):
                    file = file[nn:nn1]
                    found = False
                    for match in (re.finditer(self.buildregex(), file, flags=re.IGNORECASE)):
                        if "key" in match.group().lower():
                            continue
                        n2 = match.start()
                        file2 = file[n2:n2+2000]
                        s = soup(file2, 'lxml')
                        # print(s)
                        ps = s.findAll(re.compile(r'p|div'))
                        # print(ps)
                        # print(match.group())
                        for data in ps[1:]:
                            if not data.text or len(data.text) <= 30:
                                continue
                            else:
                                # print(match.group())
                                found = True
                                employees = data.text
                                break
                    if found:
                        break
            except Exception as e:
                print(p + "/" + t)
                print(e)
                employees = ""
            employees = employees.replace("\n", " ").replace("\t", " ").strip()
            if not employees:
                employees = ""
            # print(employees)
            employee_writer.writerow([cik, faod, cpor, employees])

if __name__ == '__main__':
    p = Parser(INPUT_FOLDER_NAME)
    # p.buildregex()
    p.parse()
    # p.extract('17485', '0000950123-09-071527.txt')
    # p.extract('9892', '0001193125-09-039210.txt')
    # p.extract('4962', '0000004962-18-000032.txt')
    # p.extract('17843', '0001193125-08-183570.txt')
    # p.extract('19745', '0000950123-10-022158.txt')
    # p.extract('19745', '0000019745-07-000004.txt')
    # p.extract('7623', '0001437749-15-001434.txt')